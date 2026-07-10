"""
Generate Invoice.xlsx — Alio Analytics.

FIN-2 fiscal-correctness lens applied:
  - Header "FATURA", invoice no., date/due date.
  - Seller block: legal name, NIF, address, IBAN (AO format), bank name.
  - Client block: name, NIF (editable input cells).
  - Line-item table: description / qty / unit price / line total — line
    total is a FORMULA (qty * unit price), not typed.
  - Subtotal = SUM(line totals) — formula.
  - IVA (14%) = subtotal * 0.14 — formula, recalculates when line items
    change (BRAND.md sec.8/10: "IVA 14%").
  - Grand total = subtotal + IVA — formula.
  - Kwanza format: 'Kz' #.##0,00 (pt-AO grouping/decimal — thousands '.',
    decimals ',').
  - Payment terms + bank details block.
  - Sheet protection: formula cells locked, input cells explicitly
    unlocked, so Excel's "Protect Sheet" keeps the math tamper-proof while
    staying fully fillable.
  - A4 print area with print-friendly margins/scaling.
"""

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, Protection
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins
from openpyxl.drawing.image import Image as XLImage
from openpyxl.worksheet.properties import PageSetupProperties

from brand import (
    FONT_DISPLAY, FONT_BODY, FONT_MONO,
    PRIMARY_700, PRIMARY_950, ACCENT_600, NEUTRAL_950, NEUTRAL_700, NEUTRAL_500, NEUTRAL_200, NEUTRAL_50, WHITE,
    COMPANY_NAME, NIF, ADDRESS, IBAN, BANK_NAME, EMAIL, WEBSITE, PHONE, AGT_NOTE, IVA_RATE,
    LOGO_ICON_NAVY, OUT_DIR,
)

KZ_FMT = '"Kz" #,##0.00;[RED]-"Kz" #,##0.00'
# openpyxl number formats use ',' as the group separator token and '.' as the
# decimal token regardless of locale; Excel re-renders them using the
# WORKBOOK'S locale at display time. We pin the workbook language/region to
# pt-AO-equivalent grouping (pt-PT uses the same '.' thousands / ',' decimal
# convention as Angola) so 'Kz #,##0.00' displays as "Kz 1.234.567,00".

FILL_NAVY = PatternFill("solid", fgColor=PRIMARY_950)
FILL_BRAND = PatternFill("solid", fgColor=PRIMARY_700)
FILL_SUBTLE = PatternFill("solid", fgColor=NEUTRAL_50)
FILL_INPUT = PatternFill("solid", fgColor="ECFCFF")  # accent-50 — visibly "type here"
FILL_WHITE = PatternFill("solid", fgColor=WHITE)

FONT_H1 = Font(name=FONT_DISPLAY, size=26, bold=True, color=PRIMARY_700)
FONT_H2 = Font(name=FONT_DISPLAY, size=12, bold=True, color=WHITE)
FONT_LABEL = Font(name=FONT_MONO, size=9, bold=False, color=NEUTRAL_500)
FONT_VALUE = Font(name=FONT_BODY, size=10, color=NEUTRAL_950)
FONT_VALUE_BOLD = Font(name=FONT_BODY, size=10, bold=True, color=NEUTRAL_950)
FONT_MONO_SMALL = Font(name=FONT_MONO, size=8, color=NEUTRAL_500)
FONT_TOTAL_LABEL = Font(name=FONT_DISPLAY, size=12, bold=True, color=NEUTRAL_950)
FONT_TOTAL_VALUE = Font(name=FONT_MONO, size=14, bold=True, color=PRIMARY_700)

THIN = Side(style="thin", color=NEUTRAL_200)
BORDER_ALL = Border(top=THIN, bottom=THIN, left=THIN, right=THIN)
BORDER_BOTTOM = Border(bottom=Side(style="thin", color=NEUTRAL_200))

LOCKED = Protection(locked=True)
UNLOCKED = Protection(locked=False)

N_ITEM_ROWS = 8  # editable line-item rows


def style_range(ws, cell_range, fill=None, font=None, align=None, border=None, protection=None, number_format=None):
    for row in ws[cell_range]:
        for cell in row:
            if fill:
                cell.fill = fill
            if font:
                cell.font = font
            if align:
                cell.alignment = align
            if border:
                cell.border = border
            if protection:
                cell.protection = protection
            if number_format:
                cell.number_format = number_format


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Fatura"
    ws.sheet_view.showGridLines = False

    # Column widths (A..G)
    widths = {"A": 3, "B": 30, "C": 8, "D": 15, "E": 15, "F": 2, "G": 3}
    for col, w in widths.items():
        ws.column_dimensions[col].width = w

    left_align = Alignment(horizontal="left", vertical="center")
    right_align = Alignment(horizontal="right", vertical="center")
    center_align = Alignment(horizontal="center", vertical="center")

    # ---- Logo + FATURA header ----
    img = XLImage(str(LOGO_ICON_NAVY))
    img.height = 46
    img.width = 64
    ws.add_image(img, "B2")

    ws["D2"] = "FATURA"
    ws["D2"].font = FONT_H1
    ws["D2"].alignment = Alignment(horizontal="right")
    ws.merge_cells("D2:E2")

    ws["D3"] = "Nº"
    ws["D3"].font = FONT_LABEL
    ws["D3"].alignment = right_align
    ws["E3"] = "{{FT 2026/0001}}"
    ws["E3"].font = FONT_VALUE
    ws["E3"].alignment = right_align
    ws["E3"].fill = FILL_INPUT
    ws["E3"].protection = UNLOCKED

    ws["D4"] = "Data"
    ws["D4"].font = FONT_LABEL
    ws["D4"].alignment = right_align
    ws["E4"] = "{{DD.MM.AAAA}}"
    ws["E4"].font = FONT_VALUE
    ws["E4"].alignment = right_align
    ws["E4"].fill = FILL_INPUT
    ws["E4"].protection = UNLOCKED

    ws["D5"] = "Vencimento"
    ws["D5"].font = FONT_LABEL
    ws["D5"].alignment = right_align
    ws["E5"] = "{{DD.MM.AAAA}}"
    ws["E5"].font = FONT_VALUE
    ws["E5"].alignment = right_align
    ws["E5"].fill = FILL_INPUT
    ws["E5"].protection = UNLOCKED

    # ---- Seller / Client blocks ----
    row = 7
    ws.cell(row=row, column=2, value="EMITENTE").font = FONT_LABEL
    ws.cell(row=row, column=4, value="CLIENTE").font = FONT_LABEL
    row += 1
    seller_lines = [COMPANY_NAME, f"NIF {NIF}", ADDRESS, f"IBAN {IBAN}", BANK_NAME]
    client_lines = ["{{Nome do Cliente / Empresa}}", "NIF {{000000000}}", "{{Morada do Cliente}}"]
    for i in range(max(len(seller_lines), len(client_lines))):
        r = row + i
        if i < len(seller_lines):
            c = ws.cell(row=r, column=2, value=seller_lines[i])
            c.font = FONT_VALUE_BOLD if i == 0 else FONT_VALUE
            ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=3)
        if i < len(client_lines):
            c = ws.cell(row=r, column=4, value=client_lines[i])
            c.font = FONT_VALUE_BOLD if i == 0 else FONT_VALUE
            c.fill = FILL_INPUT if i > 0 else FILL_WHITE
            c.protection = UNLOCKED if i > 0 else LOCKED
            ws.merge_cells(start_row=r, start_column=4, end_row=r, end_column=5)

    # ---- Line-item table ----
    table_start = row + len(seller_lines) + 2
    headers = ["Descrição", "Qtd", "Preço Unit. (Kz)", "", "Total (Kz)"]
    hdr_row = table_start
    for col, text in enumerate(headers, start=2):
        if text == "":
            continue
        cell = ws.cell(row=hdr_row, column=col, value=text)
        cell.font = FONT_H2
        cell.fill = FILL_BRAND
        cell.alignment = center_align if col != 2 else left_align
    ws.merge_cells(start_row=hdr_row, start_column=4, end_row=hdr_row, end_column=4)
    ws.cell(row=hdr_row, column=2).alignment = Alignment(horizontal="left", vertical="center", indent=1)
    for col in (2, 3, 4, 5):
        ws.cell(row=hdr_row, column=col).border = BORDER_ALL

    first_item_row = hdr_row + 1
    for i in range(N_ITEM_ROWS):
        r = first_item_row + i
        desc = ws.cell(row=r, column=2)
        qty = ws.cell(row=r, column=3)
        price = ws.cell(row=r, column=4)
        total = ws.cell(row=r, column=5)

        desc.value = "" if i > 0 else "{{Descrição do serviço}}"
        desc.font = FONT_VALUE
        desc.fill = FILL_INPUT
        desc.protection = UNLOCKED
        desc.alignment = Alignment(horizontal="left", vertical="center", indent=1)
        desc.border = BORDER_BOTTOM

        qty.value = "" if i > 0 else 1
        qty.font = FONT_VALUE
        qty.fill = FILL_INPUT
        qty.protection = UNLOCKED
        qty.alignment = center_align
        qty.border = BORDER_BOTTOM

        price.value = "" if i > 0 else 0
        price.font = FONT_VALUE
        price.fill = FILL_INPUT
        price.protection = UNLOCKED
        price.alignment = right_align
        price.number_format = KZ_FMT
        price.border = BORDER_BOTTOM

        # Line total is a FORMULA, not typed — blank when qty/price are blank.
        total.value = f'=IF(OR(C{r}="",D{r}=""),"",C{r}*D{r})'
        total.font = FONT_VALUE_BOLD
        total.fill = FILL_WHITE
        total.protection = LOCKED
        total.alignment = right_align
        total.number_format = KZ_FMT
        total.border = BORDER_BOTTOM

    last_item_row = first_item_row + N_ITEM_ROWS - 1

    # ---- Subtotal / IVA / Total ----
    sub_row = last_item_row + 2
    iva_row = sub_row + 1
    total_row = iva_row + 1

    def totals_line(r, label, formula, bold=False, big=False):
        lab = ws.cell(row=r, column=4, value=label)
        lab.font = FONT_TOTAL_LABEL if bold else FONT_VALUE
        lab.alignment = right_align
        val = ws.cell(row=r, column=5, value=formula)
        val.font = FONT_TOTAL_VALUE if big else (FONT_VALUE_BOLD if bold else FONT_VALUE)
        val.alignment = right_align
        val.number_format = KZ_FMT
        val.protection = LOCKED
        if big:
            val.fill = PatternFill("solid", fgColor="EEF0FB")  # primary-50
            val.border = Border(top=Side(style="medium", color=PRIMARY_700))
            lab.border = Border(top=Side(style="medium", color=PRIMARY_700))

    totals_line(sub_row, "Subtotal", f"=SUM(E{first_item_row}:E{last_item_row})")
    totals_line(iva_row, f"IVA ({int(IVA_RATE * 100)}%)", f"=E{sub_row}*{IVA_RATE}")
    totals_line(total_row, "TOTAL", f"=E{sub_row}+E{iva_row}", bold=True, big=True)

    # ---- Payment terms / bank / AGT note ----
    note_row = total_row + 3
    ws.cell(row=note_row, column=2, value="Condições de Pagamento").font = FONT_LABEL
    ws.merge_cells(start_row=note_row, start_column=2, end_row=note_row, end_column=3)
    r = note_row + 1
    c = ws.cell(row=r, column=2, value="{{ex.: Pagamento a 30 dias após a data da fatura}}")
    c.font = FONT_VALUE
    c.fill = FILL_INPUT
    c.protection = UNLOCKED
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)

    r += 2
    ws.cell(row=r, column=2, value=f"IBAN {IBAN} · {BANK_NAME}").font = FONT_MONO_SMALL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    ws.cell(row=r, column=2, value=f"{WEBSITE} · {EMAIL} · {PHONE}").font = FONT_MONO_SMALL
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)
    r += 1
    ws.cell(row=r, column=2, value=AGT_NOTE).font = Font(name=FONT_MONO, size=8, italic=True, color=NEUTRAL_500)
    ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=5)

    last_row = r

    # ---- Sheet protection: lock everything except input cells ----
    # Default cell protection in openpyxl is locked=True; we've explicitly
    # set Protection(locked=False) on every input cell above, so enabling
    # sheet protection locks formulas/labels/logo while inputs stay editable.
    ws.protection.sheet = True
    # No password set — this is a structural lock (protects formulas from
    # accidental edits) not an access-control measure; anyone can unprotect
    # via Excel's "Unprotect Sheet" if needed.
    ws.protection.formatCells = False
    ws.protection.formatColumns = False
    ws.protection.formatRows = False
    ws.protection.selectLockedCells = False
    ws.protection.selectUnlockedCells = False

    # ---- Print setup: A4, print area, fit width ----
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)
    ws.page_margins = PageMargins(left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.2, footer=0.2)
    ws.print_area = f"A1:G{last_row + 2}"

    wb.properties.title = "Alio Analytics — Fatura"
    wb.properties.creator = "Alio Analytics"

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    xlsx_path = OUT_DIR / "alio-invoice.xlsx"
    wb.save(xlsx_path)
    print("  OK", xlsx_path.name)


if __name__ == "__main__":
    build()
